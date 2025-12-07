import openpyxl
from datetime import datetime


class ExcelValidator:
    """Service for validating Excel files"""
    
    def validate(self, file_path, rules):
        """
        Validate an Excel file against a set of rules
        
        Args:
            file_path: Path to the Excel file
            rules: Dictionary of validation rules
            
        Returns:
            Dictionary with 'passed' (bool) and 'errors' (list)
        """
        errors = []
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
        except Exception as e:
            return {
                'passed': False,
                'errors': [f"Failed to open Excel file: {str(e)}"]
            }
        
        if 'required_sheets' in rules:
            errors.extend(self._validate_required_sheets(wb, rules['required_sheets']))
        
        if 'required_columns' in rules:
            errors.extend(self._validate_required_columns(wb, rules['required_columns']))
        
        if 'formula_sheets' in rules:
            errors.extend(self._validate_formulas(wb, rules['formula_sheets']))
        
        if 'data_validations' in rules:
            errors.extend(self._validate_data_ranges(wb, rules['data_validations']))
        
        errors.extend(self._check_circular_references(wb))
        
        wb.close()
        
        return {
            'passed': len(errors) == 0,
            'errors': errors
        }
    
    def _validate_required_sheets(self, workbook, required_sheets):
        """Check that all required sheets exist"""
        errors = []
        for sheet_name in required_sheets:
            if sheet_name not in workbook.sheetnames:
                errors.append(f"Missing required sheet: {sheet_name}")
        return errors
    
    def _validate_required_columns(self, workbook, required_columns):
        """Check that required columns exist in specified sheets"""
        errors = []
        for sheet_name, columns in required_columns.items():
            if sheet_name not in workbook.sheetnames:
                continue
            
            sheet = workbook[sheet_name]
            if sheet.max_row < 1:
                errors.append(f"Sheet '{sheet_name}' is empty")
                continue
            
            headers = [str(cell.value) for cell in sheet[1] if cell.value]
            
            for required_col in columns:
                if required_col not in headers:
                    errors.append(f"Missing required column '{required_col}' in sheet '{sheet_name}'")
        return errors
    
    def _validate_formulas(self, workbook, formula_sheets):
        """Check for broken formulas"""
        errors = []
        for sheet_name in formula_sheets:
            if sheet_name not in workbook.sheetnames:
                continue
            
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        if any(err in formula for err in ['#REF!', '#NAME?', '#VALUE!', '#DIV/0!', '#N/A']):
                            errors.append(f"Broken formula in {sheet_name}!{cell.coordinate}: {formula}")
        return errors
    
    def _validate_data_ranges(self, workbook, data_validations):
        """Validate that data falls within specified ranges"""
        errors = []
        for range_spec, validation_rule in data_validations.items():
            try:
                sheet_name, cell_range = range_spec.split('!')
                if sheet_name not in workbook.sheetnames:
                    continue
                
                sheet = workbook[sheet_name]
                validation_type = validation_rule.get('type')
                
                if validation_type == 'range':
                    min_val = validation_rule.get('min')
                    max_val = validation_rule.get('max')
                    
                    if ':' in cell_range and len(cell_range) == 3:
                        col = cell_range[0]
                        cells = sheet[col]
                    else:
                        cells = sheet[cell_range]
                    
                    for cell in cells:
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            if min_val is not None and cell.value < min_val:
                                errors.append(f"Value {cell.value} in {sheet_name}!{cell.coordinate} is below minimum {min_val}")
                            if max_val is not None and cell.value > max_val:
                                errors.append(f"Value {cell.value} in {sheet_name}!{cell.coordinate} is above maximum {max_val}")
            except Exception as e:
                errors.append(f"Error validating range {range_spec}: {str(e)}")
        return errors
    
    def _check_circular_references(self, workbook):
        """Basic check for circular references"""
        errors = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value).upper()
                        cell_ref = cell.coordinate
                        if cell_ref in formula:
                            errors.append(f"Possible circular reference in {sheet.title}!{cell_ref}")
        return errors